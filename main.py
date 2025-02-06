import streamlit as st
import pandas as pd
from datetime import datetime
import os
from utils import validate_currency, validate_date, save_to_excel, delete_row_from_excel
import tempfile
import openpyxl

# Configuração da página
st.set_page_config(
    page_title="Gerenciador de Romaneio",
    page_icon="📋",
    layout="wide"  # Alterado para "wide" para melhor uso do espaço
)

# Custom CSS para melhorar o layout
st.markdown("""
<style>
/* Estilização personalizada, se necessário */
</style>
""", unsafe_allow_html=True)

def initialize_excel_file(data):
    """Inicializa um novo arquivo Excel com estrutura básica"""
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = data.strftime('%d_%m_%Y')
    return wb

def read_sheet_data(file_path, sheet_name):
    """Lê dados de uma aba do Excel e retorna como DataFrame"""
    try:
        wb = openpyxl.load_workbook(file_path)
        # Verifica se a aba existe no arquivo. Se não, retorna None.
        if sheet_name not in wb.sheetnames:
            return None
        # Lê os dados, pulando a primeira linha (cabeçalho).
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=None, skiprows=1)
        if not df.empty:
            return df
    except Exception as e:
        # Ignora erros ao abrir o arquivo ou ler a aba
        return None
    return None

def main():
    # Lista de cidades disponíveis
    CIDADES = ["Paulínia", "Monte Mor", "Santo Antônio de Posse"]

    # Inicialização do estado da sessão
    if 'step' not in st.session_state:
        st.session_state.step = 1
    if 'current_file' not in st.session_state:
        st.session_state.current_file = None
    if 'show_download' not in st.session_state:
        st.session_state.show_download = False
    if 'cidade' not in st.session_state:
        st.session_state.cidade = CIDADES[0]  # Default para a primeira cidade
    if 'data' not in st.session_state:
        st.session_state.data = datetime.now().date()  # Data atual
    if 'current_sheet' not in st.session_state:
        st.session_state.current_sheet = None

    # Etapa 1: Informações Iniciais
    if st.session_state.step == 1:
        st.subheader("📝 Informações Iniciais")

        # Upload de planilha existente
        uploaded_file = st.file_uploader("📂 Carregar Romaneio Existente", type=['xlsx'])
        if uploaded_file:
            try:
                # Salvar o arquivo temporário
                temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
                temp_file.write(uploaded_file.getvalue())
                temp_file.close()
                st.session_state.current_file = temp_file.name

                # Ler os dados do arquivo
                wb = openpyxl.load_workbook(temp_file.name)
                sheet_name = wb.sheetnames[0]
                ws = wb[sheet_name]

                # Atualizar o estado da sessão com os dados do arquivo
                st.session_state.cidade = ws['A1'].value if ws['A1'].value in CIDADES else CIDADES[0]
                st.session_state.current_sheet = sheet_name
                try:
                    st.session_state.data = datetime.strptime(sheet_name, '%d_%m_%Y').date()
                except:
                    st.session_state.data = datetime.now().date()

                # Avançar para a próxima etapa
                st.session_state.step = 2
                st.rerun()
            except Exception as e:
                st.error(f"❌ Erro ao carregar a planilha: {e}")

        st.markdown("---")
        st.subheader("➕ Criar Novo Romaneio")
        with st.form("initial_form"):
            cidade = st.selectbox("Cidade", CIDADES, index=CIDADES.index(st.session_state.cidade))
            data_romaneio = st.date_input("Data do Romaneio", value=st.session_state.data, format="DD/MM/YYYY")
            submitted = st.form_submit_button("Criar Romaneio")

            if submitted:
                # Resetar o estado da sessão para criar um novo romaneio
                st.session_state.cidade = cidade
                st.session_state.data = data_romaneio
                st.session_state.current_file = None
                st.session_state.current_sheet = None
                st.session_state.show_download = False

                # Criar um novo arquivo Excel
                temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
                temp_file.close()
                wb = initialize_excel_file(data_romaneio)
                wb.save(temp_file.name)
                st.session_state.current_file = temp_file.name
                st.session_state.current_sheet = data_romaneio.strftime('%d_%m_%Y')
                st.session_state.step = 2
                st.rerun()

    # Etapa 2: Detalhes do Romaneio
    elif st.session_state.step == 2:
        st.subheader(f"📋 Romaneio - {st.session_state.cidade}")
        nova_data = st.date_input("📅 Data do Romaneio", value=st.session_state.data, format="DD/MM/YYYY")
        if nova_data != st.session_state.data:
            st.session_state.data = nova_data
            st.session_state.current_sheet = nova_data.strftime('%d_%m_%Y')
            st.rerun()

        if 'widget_key' not in st.session_state:
            st.session_state.widget_key = 0

        # Dividir a tela em duas colunas
        col_form, col_items = st.columns([1, 2])  # Proporção ajustada para melhor uso do espaço

        # Coluna Esquerda - Formulário
        with col_form:
    with st.expander("➕ Adicionar Item", expanded=True):
        with st.form(key=f"romaneio_form_{st.session_state.widget_key}"):
            numero_pedido = st.text_input(
                "Número do Pedido",
                placeholder="Digite o número do pedido",
                max_chars=9,
                key=f"numero_pedido_{st.session_state.widget_key}"
            )
            if numero_pedido and not numero_pedido.isdigit():
                st.error("O número do pedido deve conter apenas números.")
                return

            revendedor = st.text_input(
                "Nome do Revendedor",
                placeholder="Digite o nome do revendedor",
                key=f"revendedor_{st.session_state.widget_key}"
            )

            payment_options = ["Dinheiro", "Cartão", "Boleto"]
            pagamento = st.selectbox(
                "💳 Forma de Pagamento",
                payment_options,
                key=f"pagamento_{st.session_state.widget_key}"
            )

            valor = st.text_input(
                "💰 Valor a Pagar (R$)",
                placeholder="0,00",
                key=f"valor_{st.session_state.widget_key}"
            )

            # Botões alinhados horizontalmente
            col1, col2, col3 = st.columns([1, 1, 1])
            with col1:
                submitted_add = st.form_submit_button("➕ Adicionar")
            with col2:
                submitted_save = st.form_submit_button("💾 Salvar")
            with col3:
                submitted_home = st.form_submit_button("🔄 Tela Inicial")

            # Lógica para cada botão
            if submitted_add:
                if not numero_pedido:
                    st.error("Por favor, preencha o número do pedido.")
                    return
                if len(numero_pedido) < 9:
                    st.error("O número do pedido deve ter 9 dígitos.")
                    return
                if not revendedor:
                    st.error("Por favor, preencha o nome do revendedor.")
                    return
                revendedor = revendedor.upper()
                valor_float, error = validate_currency(valor)
                if error:
                    st.error(error)
                    return

                initial_data = [st.session_state.cidade, nova_data.strftime('%d/%m/%Y')]
                details_data = [numero_pedido, revendedor, pagamento, f"R$ {valor_float:.2f}"]
                success, message = save_to_excel(
                    [initial_data, details_data], 
                    st.session_state.current_file,
                    st.session_state.current_sheet,
                    append_mode=True
                )
                if success:
                    st.success("✅ Item adicionado com sucesso!")
                    st.session_state.widget_key += 1
                    st.rerun()
                else:
                    st.error(f"❌ Erro ao salvar: {message}")

            elif submitted_save:
                st.success("✅ Romaneio salvo com sucesso!")
                st.session_state.show_download = True

            elif submitted_home:
                # Resetar completamente o estado da sessão
                st.session_state.step = 1
                st.session_state.current_file = None
                st.session_state.current_sheet = None
                st.session_state.show_download = False
                st.session_state.widget_key = 0
                st.rerun()

        # Coluna Direita - Itens Adicionados
        with col_items:
            with st.expander("📋 Itens Adicionados", expanded=True):
                df = read_sheet_data(st.session_state.current_file, st.session_state.current_sheet)
                if df is not None and not df.empty:
                    # Remove linhas vazias
                    df = df.dropna(how='all')
                    for i, row in df.iterrows():
                        cols = st.columns([4, 1])  # Proporção ajustada para melhor uso do espaço
                        with cols[0]:
                            # Exibe a tabela sem cabeçalho
                            st.write(pd.DataFrame([row]).to_html(index=False, header=False), unsafe_allow_html=True)
                        with cols[1]:
                            if st.button(f"❌", key=f"delete_{i}", use_container_width=True):
                                success, message = delete_row_from_excel(
                                    st.session_state.current_file,
                                    st.session_state.current_sheet,
                                    i
                                )
                                if success:
                                    st.success(f"✅ Linha {i + 1} excluída com sucesso!")
                                    st.rerun()
                                else:
                                    st.error(f"❌ Erro ao excluir linha {i + 1}: {message}")
                else:
                    st.info("ℹ️ Nenhum item adicionado ainda.")

        # Botão de Download fora do formulário
        if st.session_state.show_download:
            with open(st.session_state.current_file, 'rb') as f:
                st.download_button(
                    label="📥 Baixar Planilha",
                    data=f,
                    file_name=f"Romaneio_{st.session_state.cidade}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

if __name__ == "__main__":
    main()
