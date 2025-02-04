import pandas as pd
from datetime import datetime
import openpyxl
from pathlib import Path
import streamlit as st
import re

def validate_currency(value):
    """Validate and format currency value."""
    try:
        # Remove currency symbols and spaces, replace comma with dot
        cleaned = re.sub(r'[R$\s.]', '', value).replace(',', '.')
        float_value = float(cleaned)
        if float_value < 0:
            return None, "O valor não pode ser negativo"
        return float_value, None
    except ValueError:
        return None, "Valor inválido"

def validate_date(date_str):
    """Validate date string."""
    try:
        datetime.strptime(date_str, '%Y-%m-%d')
        return True, None
    except ValueError:
        return False, "Data inválida"

def create_or_load_excel(file_path):
    """Create new Excel file or load existing one."""
    try:
        if Path(file_path).exists():
            try:
                return openpyxl.load_workbook(file_path)
            except:
                # If file exists but can't be loaded, create new
                wb = openpyxl.Workbook()
                wb.save(file_path)
                return wb
        else:
            # Create new workbook
            wb = openpyxl.Workbook()
            wb.save(file_path)
            return wb
    except Exception as e:
        raise Exception(f"Erro ao criar/carregar arquivo: {str(e)}")

def initialize_excel_file():
    """Initialize a new Excel file with basic structure"""
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = datetime.now().strftime('%d_%m_%Y')
    return wb

def save_to_excel(data, file_path, sheet_name=None, append_mode=False):
    """Save data to Excel file. Data should be a list containing [initial_data, details_data]"""
    try:
        wb = create_or_load_excel(file_path)

        if sheet_name is None:
            sheet_name = datetime.now().strftime('%d_%m_%Y')

        # Get or create sheet
        if sheet_name in wb.sheetnames:
            if append_mode:
                ws = wb[sheet_name]
            else:
                # Ensure unique sheet name if not appending
                base_name = sheet_name
                counter = 1
                while sheet_name in wb.sheetnames:
                    sheet_name = f"{base_name}_{counter}"
                    counter += 1
                ws = wb.create_sheet(sheet_name)
        else:
            ws = wb.create_sheet(sheet_name)

        # If this is first entry in sheet or not appending, write initial data
        if ws.max_row <= 1 or not append_mode:
            for col, value in enumerate(data[0], 1):
                ws.cell(row=1, column=col, value=value)
            next_row = 2
        else:
            # Find next empty row for details
            next_row = ws.max_row + 1

        # Write details data
        for col, value in enumerate(data[1], 1):
            ws.cell(row=next_row, column=col, value=value)

        try:
            wb.save(file_path)
            return True, sheet_name
        except Exception as e:
            return False, f"Erro ao salvar arquivo: {str(e)}"

    except Exception as e:
        return False, f"Erro ao processar planilha: {str(e)}"